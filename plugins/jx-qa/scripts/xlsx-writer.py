#!/usr/bin/env python3
"""
xlsx-writer.py — Pinned helper for openpyxl operations.

Subcommands:
  fork <src> <dst>        Byte-copy src xlsx to dst (shutil.copy2)
  append <xlsx> <json>    Append rows from JSON array to xlsx
  verify <xlsx>           Read xlsx and print row count + headers
  read <xlsx>             Read xlsx and print all rows as JSON

All file paths are validated: must exist (for read ops), must be .xlsx,
must not contain shell metacharacters.
"""

import sys
import os
import re
import json
import shutil

SHELL_META = re.compile(r'[;|&`$(){}!\\\n\r]')
ALLOWED_EXT = '.xlsx'


def validate_path(path, must_exist=True):
    if SHELL_META.search(path):
        print(f"ERROR: path contains shell metacharacters: {path!r}", file=sys.stderr)
        sys.exit(1)
    if not path.lower().endswith(ALLOWED_EXT):
        print(f"ERROR: only .xlsx files supported, got: {path!r}", file=sys.stderr)
        sys.exit(1)
    if must_exist and not os.path.isfile(path):
        print(f"ERROR: file not found: {path!r}", file=sys.stderr)
        sys.exit(1)


def cmd_fork(args):
    if len(args) != 2:
        print("Usage: xlsx-writer.py fork <src.xlsx> <dst.xlsx>", file=sys.stderr)
        sys.exit(1)
    src, dst = args
    validate_path(src, must_exist=True)
    validate_path(dst, must_exist=False)
    shutil.copy2(src, dst)
    print(f"Forked: {src} -> {dst}")


def cmd_append(args):
    if len(args) != 2:
        print("Usage: xlsx-writer.py append <file.xlsx> <rows.json>", file=sys.stderr)
        sys.exit(1)
    xlsx_path, json_path = args
    validate_path(xlsx_path, must_exist=True)

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    with open(json_path, 'r') as f:
        rows = json.load(f)

    if not isinstance(rows, list):
        print("ERROR: JSON must be an array of arrays", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(xlsx_path)
    print(f"Appended {len(rows)} rows to {xlsx_path}")


def cmd_verify(args):
    if len(args) != 1:
        print("Usage: xlsx-writer.py verify <file.xlsx>", file=sys.stderr)
        sys.exit(1)
    xlsx_path = args[0]
    validate_path(xlsx_path, must_exist=True)

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    print(f"Sheet: {ws.title}")
    print(f"Headers: {headers}")
    print(f"Total rows: {ws.max_row}")
    wb.close()


def cmd_read(args):
    if len(args) != 1:
        print("Usage: xlsx-writer.py read <file.xlsx>", file=sys.stderr)
        sys.exit(1)
    xlsx_path = args[0]
    validate_path(xlsx_path, must_exist=True)

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    print(json.dumps(rows, default=str))


COMMANDS = {
    'fork': cmd_fork,
    'append': cmd_append,
    'verify': cmd_verify,
    'read': cmd_read,
}


def main():
    if len(sys.argv) < 2 or sys.argv[1] not in COMMANDS:
        print(f"Usage: xlsx-writer.py <{'|'.join(COMMANDS)}> [args...]", file=sys.stderr)
        sys.exit(1)
    COMMANDS[sys.argv[1]](sys.argv[2:])


if __name__ == '__main__':
    main()
